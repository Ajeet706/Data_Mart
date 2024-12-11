# queryapp/views.py
import pandas as pd
from django.http import JsonResponse, HttpResponse
from django.http import HttpResponseRedirect
from django.shortcuts import get_object_or_404
from .models import SQLQuery
import cx_Oracle
from django.shortcuts import render,redirect
from django.db import connection
from django.views.decorators.csrf import csrf_exempt
import tempfile
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
from io import BytesIO
import os
from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required


@login_required
@permission_required('queryapp.view_sqlquery', raise_exception=True)
def query_list(request):
    queries = SQLQuery.objects.all()
    return render(request, 'queryapp/query_list.html', {'queries': queries})

@login_required
@permission_required('queryapp.download_sqlquery', raise_exception=True)
@csrf_exempt
def execute_query_and_download(request, query_id):
    # Fetch the SQLQuery object from the database
    sql_query = get_object_or_404(SQLQuery, id=query_id)

    if not sql_query:
        return HttpResponse("Query not found",status=404)
    else:
        # Execute the raw SQL query
        with connection.cursor() as cursor:
            cursor.execute(sql_query.sql_text)
            columns = [col[0] for col in cursor.description]  # Column names
            data = cursor.fetchall()  # All rows

        # Create a DataFrame and convert it to an Excel file
        df = pd.DataFrame(data, columns=columns)

        # Set up the HttpResponse to serve the Excel file
        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = f'attachment; filename="{sql_query.name}.xlsx"'
        df.to_excel(response, index=False)  # Export DataFrame to Excel in the response

    return response

def download_excel(request):
    # Create a new workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sample Data"

    # Add a header row (example headers)
    headers = ["LOANID"]
    for col_num, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)  # Make the header bold

    # Add some example data
    data = [
        [6044481],
        [6307903],
        [8237435],
    ]
    for row_num, row_data in enumerate(data, start=2):
        for col_num, cell_value in enumerate(row_data, start=1):
            sheet.cell(row=row_num, column=col_num).value = cell_value

    # Set the content type for Excel and attach the file
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="Format_file.xlsx"'

    # Save the workbook to the response
    workbook.save(response)
    return response

def upload_excel(request):
    if request.method == 'POST':
        query_id = request.POST.get('query_id')
        excel_file = request.FILES.get('excel_file')
        errors=[]

        if not excel_file:
            errors.append("No file uploaded. Please select an Excel File.")
        elif not excel_file.name.endswith('.xlsx'):
            errors.append("Invalid file format. please upload an excel file")
        else:
            try:
                workbook = load_workbook(excel_file)
                sheet = workbook.active
                if sheet.max_row==1:
                    errors.append("The excel file is empty")
                else:
                    try:
                        df=pd.read_excel(excel_file)
                        row_count = len(df)
                    except Exception as e:
                        print(f"Error reading the Excel file: (e)")
                        raise
                    if 'LOANID' not in df.columns:
                        raise ValueError("The 'LOANID' column is not present in the excel file.")
                    df.dropna(subset=['LOANID'],inplace =True)
                    if df.empty:
                        raise ValueError("DataFream is empty after dropping Nan values is 'LOANID' ")
                    table_name = "django_demo"
                    cursor = connection.cursor()

                    df['LOANID'] = df['LOANID'].astype(str)
                    try:
                        sql_truncate = f"truncate table {table_name}"
                        cursor.execute(sql_truncate)
                        connection.commit()
                        print(f"Table {table_name} truncated")
                    except Exception as e:
                        print(f"Error truncating table: (e)")
                        connection.rollback()

                    values = [(str(row['LOANID']).strip(),) for index, row in df.iterrows() if pd.notna(row['LOANID'])]

                    try:
                        sql = f"insert into {table_name} (LOANID) values (%s)"
                        cursor.executemany(sql,values)
                        connection.commit()
                        print(f"Inserting {len(values)} rows successfully.")
                    except Exception as e:
                        print(f"Error inserting rows: {e}")
                        connection.rollback()
            except Exception as e:
                errors.append(f"Error reading the Excel file: {str(e)}")

        if errors:
            return JsonResponse({"success":False,"errors":errors},status=400)

        download_url = f'/queryapp/download/{query_id}/'
        return JsonResponse({
            "success":True,
            "message":"File validated and uploaded successfully",
            "query_ID":query_id,
            "row_count":row_count,
            "Download_url":download_url
        })

    return JsonResponse({"success":False,"errors":["Invalid request menhod"]},status=400)

def view_page(request):
    with connection.cursor() as cursor:
        cursor.execute("SELECT name, position, office, age, start_date, salary, extra FROM employe_01")
        employees = cursor.fetchall()  # Fetches all rows
    return render(request, 'queryapp/test.html', {'employees': employees})